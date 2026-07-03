import pathlib
from io import BytesIO
from pathlib import PurePosixPath

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font

from config.s3_utils import get_s3_client_and_storage_options

EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)


def create_workbook() -> Workbook:
    """Create a new workbook and remove its default worksheet."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook


def create_cw_sheets(wb: Workbook, cw_end: int = 54) -> None:
    """
    Create one worksheet for each calendar week.

    Each worksheet contains bold column headers, and the header row
    remains visible while scrolling.
    """

    for week_num in range(1, cw_end):
        sheet_name = f'CW{week_num}'
        worksheet = wb.create_sheet(sheet_name)

        worksheet['A1'] = 'Emails'
        worksheet['B1'] = 'ToDo'

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        worksheet.freeze_panes = "A2"

def get_name_format(name: str) -> str:
    """
    Convert a name into a filename-friendly format.

    Example:
        "First Name" becomes "first_name".
    """
    return name.lower().strip().replace(" ", "_").replace(".", "")


def upload_workbook_to_s3(wb: Workbook, bucket_name: str, object_key: str) -> None:
    """Save an openpyxl workbook in memory and upload it to S3."""
    s3_client, _ = get_s3_client_and_storage_options()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    s3_client.upload_fileobj(buffer, bucket_name, object_key, ExtraArgs={'ContentType': EXCEL_CONTENT_TYPE})


def upload_dataframe_to_s3(df: pd.DataFrame, bucket_name: str, object_key: str) -> None:
    """Uploads the new dataframe to S3."""
    s3_client, _ = get_s3_client_and_storage_options()
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    buffer.seek(0)

    s3_client.upload_fileobj(
        buffer,
        bucket_name,
        object_key,
        ExtraArgs={
            "ContentType": EXCEL_CONTENT_TYPE,
        },
    )


def get_existing_todo_files(bucket_name: str, todo_folder: str) -> set[str]:
    """Return the filenames currently stored in the S3 ToDo folder."""
    s3_client, _ = get_s3_client_and_storage_options()

    folder_prefix = todo_folder.strip('/') + '/'
    paginator =   s3_client.get_paginator('list_objects_v2')

    existing_files = set()

    for page in paginator.paginate(Bucket=bucket_name, Prefix=folder_prefix):
        contents = page.get('Contents', [])
        for obj in contents:
            object_key = obj['Key']

            if object_key.endswith('/'):
                continue

            filename = PurePosixPath(object_key).name
            existing_files.add(filename)


    return existing_files


def create_todo_table(name: str, bucket_name, todo_folder: str) -> str:
    """Create and save a ToDo_workbook for the specified agent."""
    wb = create_workbook()
    create_cw_sheets(wb)

    formated_name= get_name_format(name)
    filename = f'{formated_name}_todo.xlsx'

    object_key = str(PurePosixPath(todo_folder) / filename)
    upload_workbook_to_s3(wb=wb, bucket_name=bucket_name, object_key=object_key)

    print(f"Uploaded s3://{bucket_name}/{object_key}")
    return filename


def add_user_table(
        user_names_df:
        pd.DataFrame,
        bucket_name: str,
        team_folder,
        todo_folder,
        names_file:str = 'names.xlsx'
) -> pd.DataFrame:
    """
    Create a ToDo_workbook for each agent who does not already have one.

    After creating each workbook, update the agent's todo_file value
    and save the updated names table."""

    required_columns = {'Name', 'todo_file'}
    missing_columns = required_columns - set(user_names_df.columns)

    if missing_columns:
        raise ValueError(f"Missing columns: {missing_columns}")

    user_names_df["todo_file"] = user_names_df["todo_file"].astype("string")
    existing_todo_files = get_existing_todo_files(bucket_name=bucket_name, todo_folder=todo_folder)

    for idx, row in user_names_df.iterrows():
        name = row['Name']

        if pd.isna(name) or not str(name).strip():
            print(f"Skipping row with index {idx} - no name")
            continue

        filename = f"{get_name_format(str(name))}_todo.xlsx"

        if filename in existing_todo_files:
            user_names_df.at[idx, 'todo_file'] = filename
            print(f"File {filename} already exists -> Skipping")
            continue

        created_filename = create_todo_table(name=name, bucket_name=bucket_name, todo_folder=todo_folder)
        user_names_df.at[idx, 'todo_file'] = created_filename
        existing_todo_files.add(filename)


        name = get_name_format(name)
        user_names_df.at[idx, 'todo_file'] = f'{name}_todo.xlsx'

    names_key = str(PurePosixPath(team_folder) / names_file)

    upload_dataframe_to_s3(
        df=user_names_df,
        bucket_name=bucket_name,
        object_key=names_key,
    )

    print(f"Uploaded s3://{bucket_name}/{names_key}")

    return user_names_df




